from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import pandas as pd
import os
import uuid
import time
from pathlib import Path
from datetime import datetime, timezone
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy as copy_style
from typing import Any

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET', 'devsecret')

ALLOWED_EXTENSIONS = {'xls', 'xlsx', 'csv'}
TMP_DIR = Path(app.root_path) / 'tmp'
TMP_DIR.mkdir(exist_ok=True)

# Cleanup configuration
TMP_RETENTION_SECONDS = int(os.environ.get('TMP_RETENTION_SECONDS', str(6 * 60 * 60)))  # default: 6 hours
TMP_CLEANUP_INTERVAL_SECONDS = int(os.environ.get('TMP_CLEANUP_INTERVAL_SECONDS', str(30 * 60)))  # default: 30 minutes
_LAST_TMP_CLEANUP = 0.0


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def cleanup_tmp_dir(now: float | None = None) -> int:
    """Delete files in TMP_DIR older than TMP_RETENTION_SECONDS.

    Returns number of files deleted.
    """
    ts_now = now if now is not None else time.time()
    deleted = 0
    try:
        for p in TMP_DIR.glob('merged-*.xlsx'):
            try:
                mtime = p.stat().st_mtime
            except FileNotFoundError:
                continue
            if ts_now - mtime > TMP_RETENTION_SECONDS:
                try:
                    p.unlink(missing_ok=True)
                    deleted += 1
                except Exception as e:
                    # Non-fatal; log and continue
                    app.logger.debug(f"Could not delete tmp file {p}: {e}")
    except Exception as e:
        app.logger.warning(f"tmp cleanup scan failed: {e}")
    if deleted:
        app.logger.info(f"tmp cleanup removed {deleted} old file(s)")
    return deleted


@app.before_request
def _maybe_cleanup_tmp():
    global _LAST_TMP_CLEANUP
    now = time.time()
    if now - _LAST_TMP_CLEANUP >= TMP_CLEANUP_INTERVAL_SECONDS:
        _LAST_TMP_CLEANUP = now
        cleanup_tmp_dir(now)


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Validate files
        if 'files' not in request.files:
            flash('No files part')
            return redirect(request.url)
        files = request.files.getlist('files')
        if not files or all(f.filename == '' for f in files):
            flash('No selected files')
            return redirect(request.url)

        # Output mode: 'single' (default) or 'sheets'
        output_mode = (request.form.get('output_mode') or 'single').strip()
        # Alignment mode (single-sheet only): 'position' (default) or 'names'
        align_mode = (request.form.get('align_mode') or 'position').strip()
        # Freeze panes toggle: default ON, but if unchecked the field is absent on POST
        freeze_panes = ('freeze_panes' in request.form)

        # Parse per-file labels and custom column title (single-sheet mode only)
        raw_labels = request.form.get('group_texts', '')
        labels = [l.strip() for l in raw_labels.splitlines()] if raw_labels else []
        group_col_title = (request.form.get('group_col_title') or '').strip() or 'Group'
        use_group_col = (output_mode != 'sheets') and any(bool(l) for l in labels)

        # Parse optional sheet names (separate-sheets mode only)
        raw_names = request.form.get('sheet_names', '') if output_mode == 'sheets' else ''
        requested_names = [n.strip() for n in raw_names.splitlines()] if raw_names else []

        def sanitize_sheet_names(default_bases: list[str]) -> list[str]:
            r"""Return sanitized, unique Excel sheet names.

            - limit to 31 chars
            - remove invalid chars : \\ / ? * [ ]
            - trim quotes and spaces; avoid empty by falling back to defaults
            - ensure uniqueness by appending (2), (3), ...
            """
            invalid_chars = set(':\\/?*[]')
            used: set[str] = set()
            result: list[str] = []
            for i, base in enumerate(default_bases):
                raw = (requested_names[i] if i < len(requested_names) and requested_names[i] else base)
                # remove invalid characters
                cleaned = ''.join(ch for ch in raw if ch not in invalid_chars)
                cleaned = cleaned.strip().strip("'")
                if not cleaned:
                    cleaned = base or f'Sheet{i+1}'
                # limit to 31 characters
                cleaned = cleaned[:31]
                # enforce uniqueness
                base_clean = cleaned
                n = 2
                while cleaned in used or not cleaned:
                    suffix = f' ({n})'
                    max_base_len = 31 - len(suffix)
                    trimmed = base_clean[:max_base_len].rstrip()
                    cleaned = f"{trimmed}{suffix}" if trimmed else f"Sheet{i+1}{suffix}"
                    n += 1
                used.add(cleaned)
                result.append(cleaned)
            return result

        def autosize_columns(ws) -> None:
            """Set column widths to fit the longest cell content (approx), with padding and bounds."""
            try:
                max_col = ws.max_column or 0
                max_row = ws.max_row or 0
                if max_col == 0 or max_row == 0:
                    return
                for col_idx in range(1, max_col + 1):
                    max_len = 0
                    for row_idx in range(1, max_row + 1):
                        v = ws.cell(row=row_idx, column=col_idx).value
                        if v is None:
                            continue
                        s = str(v)
                        for line in s.splitlines() or [s]:
                            if len(line) > max_len:
                                max_len = len(line)
                    # Apply padding and bounds
                    width = max_len + 2
                    if width < 10:
                        width = 10
                    if width > 60:
                        width = 60
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
            except Exception:
                # Non-fatal; ignore autosize errors
                pass

        dfs = []
        merge_ranges = []  # tuples of (start_idx, end_idx) for data rows per file in merged DF (0-based)
        running_total = 0
        first_cols: list[str] | None = None  # for position alignment

        # For separate sheets, collect (name, df, file, ext)
        per_sheet: list[tuple[str, pd.DataFrame, Any, str]] = []
        default_sheet_bases: list[str] = []

        for i, file in enumerate(files):
            raw_fn = (getattr(file, 'filename', '') or '')
            filename = secure_filename(raw_fn)
            if not (file and allowed_file(filename)):
                flash(f'File not allowed: {getattr(file, "filename", "")}')
                return redirect(request.url)

            try:
                if filename.lower().endswith('.csv'):
                    df = pd.read_csv(file.stream)
                else:
                    df = pd.read_excel(file.stream)
            except Exception as e:
                flash(f'Error reading {filename}: {e}')
                return redirect(request.url)

            n_data = len(df)
            original_cols = list(df.columns)

            if output_mode == 'single':
                if n_data > 0:
                    # Build Average row using original columns
                    avg_row: dict[str, Any] = {original_cols[0]: 'Average'}
                    for c in original_cols[1:]:
                        numeric = pd.to_numeric(df[c], errors='coerce')
                        mean_val = numeric.mean()
                        avg_row[c] = '' if pd.isna(mean_val) else mean_val
                    avg_df = pd.DataFrame([avg_row], columns=original_cols)

                    # Position-based alignment to first non-empty file's columns
                    if align_mode == 'position':
                        if first_cols is None:
                            first_cols = original_cols.copy()
                        taken = min(len(original_cols), len(first_cols))
                        pos_aligned = df.iloc[:, :taken].copy()
                        pos_aligned.columns = first_cols[:taken]
                        for extra_col in first_cols[taken:]:
                            pos_aligned[extra_col] = ''
                        df = pos_aligned[first_cols]
                        # Align avg_df to first_cols by POSITION (not by name)
                        avg_row_aligned: dict[str, Any] = {c: '' for c in first_cols}
                        for j in range(taken):
                            val = avg_df.iloc[0, j]
                            avg_row_aligned[first_cols[j]] = ('' if pd.isna(val) else val)
                        avg_df = pd.DataFrame([avg_row_aligned], columns=first_cols)

                    # Insert group column if needed (after alignment)
                    if use_group_col:
                        label = labels[i] if i < len(labels) else ''
                        df.insert(0, group_col_title, label)
                        avg_df.insert(0, group_col_title, '')
                        cols_after = [group_col_title] + (first_cols if (align_mode == 'position' and first_cols is not None) else original_cols)
                    else:
                        cols_after = (first_cols if (align_mode == 'position' and first_cols is not None) else original_cols)

                    # Record merge range for this file's data rows if it has a label
                    if use_group_col and (labels[i] if i < len(labels) else '') and n_data > 0:
                        start_idx = running_total
                        end_idx = running_total + n_data - 1
                        merge_ranges.append((start_idx, end_idx))

                    # Append Average and optional empty row (not after last file)
                    if i < len(files) - 1:
                        empty_row = {c: '' for c in cols_after}
                        empty_df = pd.DataFrame([empty_row], columns=cols_after)
                        df = pd.concat([df, avg_df, empty_df], ignore_index=True)
                    else:
                        df = pd.concat([df, avg_df], ignore_index=True)

                    running_total += len(df)
                # collect regardless
                dfs.append(df)
            else:
                # Separate sheets: write data as-is (no Average row, no group column)
                dfs.append(df)  # keep for potential uniform preview handling
                default_sheet_bases.append(Path(filename).stem or f'Sheet{i+1}')
                ext = Path(filename).suffix.lower()
                per_sheet.append((filename, df, file, ext))

        if output_mode == 'sheets':
            # Sanitize and finalize sheet names
            sheet_names = sanitize_sheet_names(default_sheet_bases)

            # Write to Excel with multiple sheets, preserving merges for .xlsx sources
            unique_name = f'merged-{uuid.uuid4().hex}.xlsx'
            out_path = TMP_DIR / unique_name
            try:
                wb_out = Workbook()
                # Remove default sheet if present and empty
                if wb_out.active and wb_out.active.max_row == 1 and wb_out.active.max_column == 1 and wb_out.active['A1'].value is None:
                    wb_out.remove(wb_out.active)

                for idx, (orig_name, df, uploaded_file, ext) in enumerate(per_sheet):
                    sheet_name = sheet_names[idx]
                    if ext == '.xlsx':
                        try:
                            # Reset stream and load source workbook
                            if hasattr(uploaded_file, 'stream'):
                                uploaded_file.stream.seek(0)
                            wb_src = load_workbook(uploaded_file.stream, data_only=True)
                            ws_src = getattr(wb_src, 'active', None)
                            ws_dest = wb_out.create_sheet(title=sheet_name)
                            if ws_src is not None and hasattr(ws_src, 'iter_rows') and hasattr(ws_src, 'merged_cells'):
                                # Copy values and styles
                                for r in ws_src.iter_rows(values_only=False):
                                    for cell in r:
                                        dcell = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
                                        try:
                                            if getattr(cell, 'font', None) is not None:
                                                dcell.font = copy_style(cell.font)
                                            if getattr(cell, 'fill', None) is not None:
                                                dcell.fill = copy_style(cell.fill)
                                            if getattr(cell, 'border', None) is not None:
                                                dcell.border = copy_style(cell.border)
                                            if getattr(cell, 'alignment', None) is not None:
                                                dcell.alignment = copy_style(cell.alignment)
                                            if getattr(cell, 'number_format', None) is not None:
                                                dcell.number_format = cell.number_format
                                            if getattr(cell, 'protection', None) is not None:
                                                dcell.protection = copy_style(cell.protection)
                                        except Exception:
                                            # Non-fatal style copy failure; continue with values
                                            pass
                                # Reapply merged ranges
                                for mcr in ws_src.merged_cells.ranges:
                                    ws_dest.merge_cells(range_string=str(mcr.coord))
                            else:
                                # Fallback to DataFrame values
                                ws_dest.append(list(df.columns))
                                for _, s in df.iterrows():
                                    ws_dest.append([s.get(c, '') for c in df.columns])
                            # Freeze first row and first column
                            if freeze_panes:
                                ws_dest.freeze_panes = 'B2'
                            # Center align all cells
                            try:
                                center = Alignment(horizontal='center', vertical='center')
                                for r in ws_dest.iter_rows(min_row=1, max_row=ws_dest.max_row, min_col=1, max_col=ws_dest.max_column):
                                    for cell in r:
                                        cell.alignment = center
                                autosize_columns(ws_dest)
                            except Exception:
                                pass
                        except Exception as e:
                            # Fallback: write values from DataFrame if anything fails
                            ws_dest = wb_out.create_sheet(title=sheet_name)
                            # Write headers
                            ws_dest.append(list(df.columns))
                            for _, s in df.iterrows():
                                ws_dest.append([s.get(c, '') for c in df.columns])
                            if freeze_panes:
                                ws_dest.freeze_panes = 'B2'
                            try:
                                center = Alignment(horizontal='center', vertical='center')
                                for r in ws_dest.iter_rows(min_row=1, max_row=ws_dest.max_row, min_col=1, max_col=ws_dest.max_column):
                                    for cell in r:
                                        cell.alignment = center
                                autosize_columns(ws_dest)
                            except Exception:
                                pass
                    else:
                        # CSV or other: write values from DataFrame
                        ws_dest = wb_out.create_sheet(title=sheet_name)
                        ws_dest.append(list(df.columns))
                        for _, s in df.iterrows():
                            ws_dest.append([s.get(c, '') for c in df.columns])
                        if freeze_panes:
                            ws_dest.freeze_panes = 'B2'
                        try:
                            center = Alignment(horizontal='center', vertical='center')
                            for r in ws_dest.iter_rows(min_row=1, max_row=ws_dest.max_row, min_col=1, max_col=ws_dest.max_column):
                                for cell in r:
                                    cell.alignment = center
                            autosize_columns(ws_dest)
                        except Exception:
                            pass

                wb_out.save(out_path)
            except Exception as e:
                flash(f'Error writing output file: {e}')
                return redirect(request.url)

            # Build previews
            sheet_previews = []
            for idx, (_, df, _, _) in enumerate(per_sheet):
                html = df.to_html(classes='table table-striped table-sm', index=False, escape=False)
                sheet_previews.append({
                    'name': sheet_names[idx],
                    'html': html,
                    'rows': len(df),
                    'cols': len(df.columns)
                })

            return render_template('index.html', sheet_previews=sheet_previews, download_filename=unique_name)

        # Single-sheet mode: merge all
        try:
            merged = pd.concat(dfs, ignore_index=True)
        except Exception as e:
            flash(f'Error merging files: {e}')
            return redirect(request.url)

        # Write to Excel
        unique_name = f'merged-{uuid.uuid4().hex}.xlsx'
        out_path = TMP_DIR / unique_name
        try:
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                merged.to_excel(writer, index=False, sheet_name='Merged')
        except Exception as e:
            flash(f'Error writing output file: {e}')
            return redirect(request.url)

        # Apply freeze panes and (optionally) merge labeled ranges in single-sheet output
        try:
            wb = load_workbook(out_path)
            ws = wb['Merged']
            # Freeze first row and first column
            if freeze_panes:
                ws.freeze_panes = 'B2'
            # Merge first column cells for labeled ranges (single-sheet only)
            if use_group_col and merge_ranges:
                for (start_idx, end_idx) in merge_ranges:
                    excel_start = start_idx + 2  # header row + 1-based indexing
                    excel_end = end_idx + 2
                    if excel_start <= excel_end:
                        ws.merge_cells(start_row=excel_start, start_column=1, end_row=excel_end, end_column=1)
            # Apply light yellow background to Average rows and merged group cells
            try:
                yellow = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD', fill_type='solid')
                # Average rows: any row containing 'Average'
                for cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    if any((cell.value == 'Average') for cell in cells):
                        for cell in cells:
                            cell.fill = yellow
                # Merged group cells (first column ranges)
                if use_group_col and merge_ranges:
                    for (start_idx, end_idx) in merge_ranges:
                        excel_start = start_idx + 2
                        excel_end = end_idx + 2
                        # Set fill on the top-left cell of the merged area
                        ws.cell(row=excel_start, column=1).fill = yellow
            except Exception:
                pass
            # Center align all cells
            try:
                center = Alignment(horizontal='center', vertical='center')
                for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in r:
                        cell.alignment = center
                autosize_columns(ws)
            except Exception:
                pass
            wb.save(out_path)
        except Exception as e:
            flash(f'Warning: could not finalize Excel formatting: {e}')

        # Render preview
        merged_html = merged.to_html(classes='table table-striped table-sm', index=False, escape=False)
        return render_template('index.html', merged_table=merged_html, download_filename=unique_name, num_rows=len(merged), num_cols=len(merged.columns))

    # GET
    return render_template('index.html')


@app.route('/download/<path:filename>')
def download(filename: str):
    # sanitize and only allow files created by this app (merged-*.xlsx)
    if not filename.startswith('merged-') or not filename.endswith('.xlsx'):
        flash('Invalid download request')
        return redirect(url_for('index'))
    file_path = TMP_DIR / filename
    if not file_path.exists():
        flash('File not found or expired')
        return redirect(url_for('index'))
    try:
        return send_file(
            file_path,
            as_attachment=True,
            download_name='merged.xlsx'
        )
    except Exception as e:
        flash(f'Error sending file: {e}')
        return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(debug=True)
